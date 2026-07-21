from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


output_directory = Path(sys.argv[1])
output_directory.mkdir(parents=True, exist_ok=True)
output_path = output_directory / "OpenDeskTW_完整試算表功能.xlsx"

workbook = Workbook()
dashboard = workbook.active
dashboard.title = "儀表板"
data = workbook.create_sheet("資料表")
settings = workbook.create_sheet("設定")
hidden = workbook.create_sheet("系統資料")
hidden.sheet_state = "hidden"

font_name = "Noto Sans CJK TC"
teal = "0F766E"
navy = "0F3D56"
light_teal = "CCFBF1"
light_gray = "F8FAFC"
thin = Side(style="thin", color="CBD5E1")
all_borders = Border(top=thin, bottom=thin, left=thin, right=thin)

dashboard.merge_cells("A1:E1")
dashboard["A1"] = "OpenDesk TW 試算表進階功能矩陣"
dashboard["A1"].font = Font(name=font_name, size=20, bold=True, color="FFFFFF")
dashboard["A1"].fill = PatternFill("solid", fgColor=navy)
dashboard["A1"].alignment = Alignment(horizontal="center", vertical="center")
dashboard.row_dimensions[1].height = 34
dashboard.append(["月份", "收入", "成本", "毛利", "達成率"])
monthly_rows = [
    ["一月", 120000, 72000, "=B3-C3", "=D3/$B$10"],
    ["二月", 138000, 79000, "=B4-C4", "=D4/$B$10"],
    ["三月", 151000, 82500, "=B5-C5", "=D5/$B$10"],
    ["四月", 146000, 81000, "=B6-C6", "=D6/$B$10"],
    ["五月", 163000, 88500, "=B7-C7", "=D7/$B$10"],
    ["六月", 176000, 94000, "=B8-C8", "=D8/$B$10"],
]
for row in monthly_rows:
    dashboard.append(row)
dashboard.append(["合計", "=SUM(B3:B8)", "=SUM(C3:C8)", "=SUM(D3:D8)", "=AVERAGE(E3:E8)"])
dashboard.append(["毛利目標", 75000, "", "跨工作表稅率", "=設定!B2"])

for cell in dashboard[2]:
    cell.font = Font(name=font_name, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=teal)
    cell.alignment = Alignment(horizontal="center")
for row in dashboard.iter_rows(min_row=3, max_row=10, min_col=1, max_col=5):
    for cell in row:
        cell.font = Font(name=font_name, size=11, bold=cell.row in (9, 10))
        cell.border = all_borders
        cell.fill = PatternFill("solid", fgColor=light_teal if cell.row in (9, 10) else "FFFFFF")
        cell.alignment = Alignment(vertical="center")
for row in range(3, 10):
    dashboard[f"B{row}"].number_format = '#,##0'
    dashboard[f"C{row}"].number_format = '#,##0'
    dashboard[f"D{row}"].number_format = '#,##0'
    dashboard[f"E{row}"].number_format = '0.0%'

dashboard.freeze_panes = "B3"
dashboard.auto_filter.ref = "A2:E9"
dashboard["B10"].comment = Comment("毛利目標可由管理者調整；用來驗證儲存格註解。", "OpenDesk TW")
dashboard.conditional_formatting.add("D3:D8", CellIsRule(operator="greaterThanOrEqual", formula=["$B$10"], fill=PatternFill("solid", fgColor="DCFCE7")))
dashboard.conditional_formatting.add("E3:E8", ColorScaleRule(start_type="min", start_color="FEE2E2", mid_type="percentile", mid_value=50, mid_color="FEF3C7", end_type="max", end_color="DCFCE7"))
status_validation = DataValidation(type="list", formula1='"草稿,審核中,已完成"', allow_blank=False)
dashboard.add_data_validation(status_validation)
dashboard["G2"] = "狀態"
dashboard["G2"].font = Font(name=font_name, bold=True, color="FFFFFF")
dashboard["G2"].fill = PatternFill("solid", fgColor=teal)
dashboard["G3"] = "審核中"
status_validation.add(dashboard["G3"])

dashboard_table = Table(displayName="MonthlyPerformance", ref="A2:E9")
dashboard_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
dashboard.add_table(dashboard_table)

chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "每月毛利"
chart.y_axis.title = "金額"
chart.x_axis.title = "月份"
chart.add_data(Reference(dashboard, min_col=4, min_row=2, max_row=8), titles_from_data=True)
chart.set_categories(Reference(dashboard, min_col=1, min_row=3, max_row=8))
chart.height = 7.8
chart.width = 13.5
dashboard.add_chart(chart, "G5")

for column, width in {"A": 14, "B": 15, "C": 15, "D": 15, "E": 14, "F": 3, "G": 16}.items():
    dashboard.column_dimensions[column].width = width
dashboard.page_setup.orientation = "landscape"
dashboard.page_setup.paperSize = dashboard.PAPERSIZE_A4
dashboard.page_setup.fitToWidth = 1
dashboard.page_setup.fitToHeight = 1
dashboard.sheet_properties.pageSetUpPr.fitToPage = True
dashboard.print_area = "A1:N22"
dashboard.print_title_rows = "1:2"
dashboard.oddHeader.center.text = "OpenDesk TW 試算表功能驗證"
dashboard.oddFooter.center.text = "第 &P 頁／共 &N 頁"
dashboard.protection.sheet = True
dashboard.protection.password = "ODTW"
dashboard["B10"].protection = Protection(locked=False)
dashboard["G3"].protection = Protection(locked=False)

data.append(["日期", "部門", "品項", "數量", "單價", "小計"])
transactions = [
    ["2026-01-05", "北區", "顧問服務", 3, 42000, "=D2*E2"],
    ["2026-02-12", "中區", "教育訓練", 5, 18000, "=D3*E3"],
    ["2026-03-18", "南區", "文件轉換", 12, 3500, "=D4*E4"],
    ["2026-04-21", "北區", "維護服務", 6, 12000, "=D5*E5"],
    ["2026-05-30", "中區", "顧問服務", 2, 46000, "=D6*E6"],
]
for row in transactions:
    data.append(row)
for row in data.iter_rows(min_row=1, max_row=6, min_col=1, max_col=6):
    for cell in row:
        cell.font = Font(name=font_name, bold=cell.row == 1, color="FFFFFF" if cell.row == 1 else "000000")
        cell.fill = PatternFill("solid", fgColor=teal if cell.row == 1 else "FFFFFF")
        cell.border = all_borders
        cell.alignment = Alignment(vertical="center")
for column, width in {"A": 16, "B": 14, "C": 20, "D": 12, "E": 14, "F": 16}.items():
    data.column_dimensions[column].width = width
data.freeze_panes = "A2"
data.auto_filter.ref = "A1:F6"
transaction_table = Table(displayName="TransactionData", ref="A1:F6")
transaction_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
data.add_table(transaction_table)
data.row_dimensions.group(4, 5, hidden=False)
data.page_setup.paperSize = data.PAPERSIZE_A4
data.page_setup.fitToWidth = 1
data.sheet_properties.pageSetUpPr.fitToPage = True

settings.append(["設定", "值"])
settings.append(["稅率", 0.05])
settings.append(["地區", "臺灣"])
settings.append(["預設字型", font_name])
for row in settings.iter_rows():
    for cell in row:
        cell.font = Font(name=font_name, bold=cell.row == 1, color="FFFFFF" if cell.row == 1 else "000000")
        cell.fill = PatternFill("solid", fgColor=teal if cell.row == 1 else light_gray)
        cell.border = all_borders
settings["B2"].number_format = "0%"
settings.column_dimensions["A"].width = 20
settings.column_dimensions["B"].width = 24

hidden.append(["鍵", "值"])
hidden.append(["產生器", "OpenDesk TW"])
hidden.append(["版本", "1.4.0"])

workbook.defined_names.add(DefinedName("TaxRate", attr_text="'設定'!$B$2"))
workbook.defined_names.add(DefinedName("MonthlyProfit", attr_text="'儀表板'!$D$3:$D$8"))
workbook.calculation.fullCalcOnLoad = True
workbook.calculation.forceFullCalc = True
workbook.calculation.calcMode = "auto"
workbook.save(output_path)

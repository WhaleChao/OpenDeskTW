import json
from pathlib import Path
import sys

from openpyxl import load_workbook

path = Path(sys.argv[1])
formulas_book = load_workbook(path, data_only=False)
values_book = load_workbook(path, data_only=True)

formula_cells = []
errors = []
for sheet in formulas_book.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append(f"{sheet.title}!{cell.coordinate}")

for sheet in values_book.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and any(
                marker in cell.value
                for marker in ["#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NUM!"]
            ):
                errors.append(f"{sheet.title}!{cell.coordinate}:{cell.value}")

summary_formula = formulas_book["總覽"]["B7"].value
summary_value = values_book["總覽"]["B7"].value
report = {
    "file": str(path.resolve()),
    "sheets": formulas_book.sheetnames,
    "formulaCells": formula_cells,
    "formulaErrors": errors,
    "summaryFormula": summary_formula,
    "summaryValue": summary_value,
    "chartCount": sum(len(sheet._charts) for sheet in formulas_book.worksheets),
    "font": formulas_book["總覽"]["A1"].font.name,
}
report["passed"] = (
    report["sheets"] == ["總覽", "資料", "設定"]
    and len(formula_cells) == 5
    and not errors
    and summary_formula == "=SUM(B3:B6)"
    and summary_value == 26
    and report["chartCount"] == 1
    and report["font"] == "Noto Sans CJK TC"
)
print(json.dumps(report, ensure_ascii=False, indent=2))
formulas_book.close()
values_book.close()
sys.exit(0 if report["passed"] else 1)

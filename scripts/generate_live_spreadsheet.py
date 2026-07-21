from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

output_dir = Path(sys.argv[1])
output_dir.mkdir(parents=True, exist_ok=True)
output_path = output_dir / "OpenDeskTW_LIVE_Sheets.xlsx"

workbook = Workbook()
summary = workbook.active
summary.title = "總覽"
data = workbook.create_sheet("資料")
settings = workbook.create_sheet("設定")

teal = "0F766E"
navy = "0F3D56"
light = "D9F3F0"
thin = Side(style="thin", color="CBD5E1")

summary.merge_cells("A1:D1")
summary["A1"] = "OpenDesk TW 試算表 LIVE 驗證"
summary["A1"].font = Font(name="Noto Sans CJK TC", size=20, bold=True, color="FFFFFF")
summary["A1"].fill = PatternFill("solid", fgColor=navy)
summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
summary.row_dimensions[1].height = 34
summary.append(["項目", "數值", "單位", "驗證內容"])
rows = [
    ["Writer 測試", 8, "項", "頁碼、頁首、表格、字型"],
    ["Sheets 測試", 7, "項", "公式、格式、圖表、頁籤"],
    ["Slides 測試", 6, "項", "版面、圖形、表格、字型"],
    ["PDF 測試", 5, "項", "匯出、頁數、搜尋文字"],
]
for row in rows:
    summary.append(row)
summary.append(["總計", "=SUM(B3:B6)", "項", "公式應為 26"])

for cell in summary[2]:
    cell.font = Font(name="Noto Sans CJK TC", bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=teal)
    cell.alignment = Alignment(horizontal="center")
for row in summary.iter_rows(min_row=3, max_row=7, min_col=1, max_col=4):
    for cell in row:
        cell.font = Font(name="Noto Sans CJK TC", size=11, bold=cell.row == 7)
        cell.fill = PatternFill("solid", fgColor=light if cell.row == 7 else "FFFFFF")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        cell.alignment = Alignment(vertical="center")
summary.freeze_panes = "A3"
summary.auto_filter.ref = "A2:D7"
for index, width in enumerate([18, 12, 10, 34], 1):
    summary.column_dimensions[get_column_letter(index)].width = width

chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "各模組驗證項目"
chart.y_axis.title = "項目數"
chart.x_axis.title = "模組"
chart.add_data(Reference(summary, min_col=2, min_row=2, max_row=6), titles_from_data=True)
chart.set_categories(Reference(summary, min_col=1, min_row=3, max_row=6))
chart.height = 8
chart.width = 14
summary.add_chart(chart, "F2")
summary.page_setup.orientation = "landscape"
summary.page_setup.paperSize = summary.PAPERSIZE_A4
summary.page_setup.fitToWidth = 1
summary.page_setup.fitToHeight = 1
summary.sheet_properties.pageSetUpPr.fitToPage = True
summary.print_area = "A1:M20"
summary.print_options.horizontalCentered = True

data.append(["日期", "模組", "測試值", "是否通過"])
for index, module in enumerate(["Writer", "Sheets", "Slides", "PDF"], 1):
    data.append([f"2026-07-{20 + index:02d}", module, index * 10, f"=C{index + 1}>0"])
for row in data.iter_rows():
    for cell in row:
        cell.font = Font(name="Noto Sans CJK TC", bold=cell.row == 1, color="FFFFFF" if cell.row == 1 else "000000")
        if cell.row == 1:
            cell.fill = PatternFill("solid", fgColor=teal)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
data.freeze_panes = "A2"
data.auto_filter.ref = "A1:D5"
data.page_setup.paperSize = data.PAPERSIZE_A4
data.page_setup.fitToWidth = 1
data.page_setup.fitToHeight = 1
data.sheet_properties.pageSetUpPr.fitToPage = True
data.print_options.horizontalCentered = True
for index, width in enumerate([16, 16, 14, 16], 1):
    data.column_dimensions[get_column_letter(index)].width = width

settings.append(["設定", "值"])
settings.append(["預設字型", "Noto Sans CJK TC"])
settings.append(["備份版本", 20])
settings.append(["保存天數", 30])
settings.append(["模式", "本機離線"])
for row in settings.iter_rows():
    for cell in row:
        cell.font = Font(name="Noto Sans CJK TC", bold=cell.row == 1, color="FFFFFF" if cell.row == 1 else "000000")
        cell.fill = PatternFill("solid", fgColor=teal if cell.row == 1 else "FFFFFF")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
settings.column_dimensions["A"].width = 20
settings.column_dimensions["B"].width = 25
settings.page_setup.paperSize = settings.PAPERSIZE_A4
settings.page_setup.fitToWidth = 1
settings.page_setup.fitToHeight = 1
settings.sheet_properties.pageSetUpPr.fitToPage = True
settings.print_options.horizontalCentered = True

workbook.calculation.fullCalcOnLoad = True
workbook.calculation.forceFullCalc = True
workbook.calculation.calcMode = "auto"
workbook.save(output_path)

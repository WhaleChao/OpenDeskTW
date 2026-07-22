from __future__ import annotations

import json
from pathlib import Path
import sys
import zipfile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(sys.argv[1])
REPORT_PATH = Path(sys.argv[2]) if len(sys.argv) > 2 else None


def package_text(path: Path, name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        return archive.read(name).decode("utf-8")


def package_names(path: Path) -> set[str]:
    with zipfile.ZipFile(path) as archive:
        return set(archive.namelist())


def check(name: str, outcome: bool, detail: str) -> dict[str, object]:
    return {"name": name, "passed": bool(outcome), "detail": detail}


def verify_docx(path: Path) -> list[dict[str, object]]:
    names = package_names(path)
    document = package_text(path, "word/document.xml")
    settings = package_text(path, "word/settings.xml")
    styles = package_text(path, "word/styles.xml")
    numbering = package_text(path, "word/numbering.xml")
    relationships = package_text(path, "word/_rels/document.xml.rels")
    checks = [
        check("DOCX 基本結構", "word/document.xml" in names and "word/styles.xml" in names, "document.xml 與 styles.xml"),
        check("DOCX 四層標題", all(f'w:styleId="Heading{level}"' in styles for level in range(1, 5)), "Heading 1–4"),
        check("DOCX 自動目錄", "TOC" in document and "1-4" in document, "TOC 欄位涵蓋 Heading 1–4"),
        check("DOCX 自動中文編號", "ideographLegalTraditional" in numbering and "〔%1、〕" in numbering, "繁體中文法律編號格式"),
        check("DOCX 頁首頁尾與頁碼", any(name.startswith("word/header") for name in names) and any(name.startswith("word/footer") for name in names) and "PAGE" in "".join(package_text(path, name) for name in names if name.startswith("word/footer") and name.endswith(".xml")), "頁首、頁尾與 PAGE 欄位"),
        check("DOCX 註腳與尾註", "word/footnotes.xml" in names and "word/endnotes.xml" in names and "footnoteReference" in document and "endnoteReference" in document, "註腳／尾註零件與參照"),
        check("DOCX 註解", "word/comments.xml" in names and "commentRangeStart" in document and "commentReference" in document, "註解零件與範圍標記"),
        check("DOCX 追蹤修訂", "<w:trackRevisions" in settings and "<w:ins" in document and "<w:del" in document, "追蹤模式、插入與刪除"),
        check("DOCX 書籤與超連結", "bookmarkStart" in document and "w:anchor=" in document and "hyperlink" in relationships, "內部書籤與外部連結"),
        check("DOCX 郵件合併欄位", "MERGEFIELD" in document and "RecipientName" in document, "SimpleMailMergeField"),
        check("DOCX 表格與分頁", "<w:tbl>" in document and "w:type=\"page\"" in document, "表格與明確分頁"),
        check("DOCX 繁中字型", "Noto Sans CJK TC" in styles and "Noto Serif CJK TC" in styles, "黑體與明體樣式"),
    ]
    return checks


def verify_xlsx(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=False)
    dashboard = workbook["儀表板"]
    names = package_names(path)
    workbook_xml = package_text(path, "xl/workbook.xml")
    dashboard_xml = package_text(path, "xl/worksheets/sheet1.xml")
    formula_cells = [cell.coordinate for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]
    checks = [
        check("XLSX 多工作表與隱藏頁籤", workbook.sheetnames == ["儀表板", "資料表", "設定", "系統資料"] and workbook["系統資料"].sheet_state == "hidden", "4 張工作表，含隱藏系統資料"),
        check("XLSX 公式與跨表參照", len(formula_cells) >= 15 and dashboard["E10"].value == "=設定!B2", f"{len(formula_cells)} 個公式"),
        check("XLSX 命名範圍", set(workbook.defined_names) >= {"TaxRate", "MonthlyProfit"}, "TaxRate、MonthlyProfit"),
        check("XLSX 格式化表格與篩選", len(dashboard.tables) == 1 and dashboard.auto_filter.ref == "A2:E9", "MonthlyPerformance 與自動篩選"),
        check("XLSX 凍結窗格", str(dashboard.freeze_panes) == "B3", "B3"),
        check("XLSX 圖表", len(dashboard._charts) == 1 and any(name.startswith("xl/charts/chart") for name in names), "1 個可編輯長條圖"),
        check("XLSX 條件格式", len(dashboard.conditional_formatting) >= 2, "儲存格規則與色階"),
        check("XLSX 資料驗證", dashboard.data_validations.count == 1 and "dataValidations" in dashboard_xml, "狀態下拉選單"),
        check("XLSX 註解", dashboard["B10"].comment is not None and any(name.startswith("xl/comments") for name in names), "目標值註解"),
        check("XLSX 工作表保護", dashboard.protection.sheet and "sheetProtection" in dashboard_xml, "受保護工作表及可編輯欄位"),
        check("XLSX 列印版面", dashboard.print_area == "'儀表板'!$A$1:$N$22" and dashboard.print_title_rows == "$1:$2", "A4 橫向、列印範圍與重複標題列"),
        check("XLSX 頁首頁尾", "oddHeader" in dashboard_xml and "oddFooter" in dashboard_xml, "列印頁首與頁碼"),
        check("XLSX 群組與字型", workbook["資料表"].row_dimensions[4].outlineLevel == 1 and dashboard["A1"].font.name == "Noto Sans CJK TC", "資料列群組及繁中字型"),
        check("XLSX 自動重算", 'calcMode="auto"' in workbook_xml and 'fullCalcOnLoad="1"' in workbook_xml, "開啟時完整重算"),
    ]
    workbook.close()
    return checks


def verify_pptx(path: Path) -> list[dict[str, object]]:
    names = package_names(path)
    slide_names = sorted(name for name in names if name.startswith("ppt/slides/slide") and name.endswith(".xml"))
    slide_xml = "".join(package_text(path, name) for name in slide_names)
    presentation = package_text(path, "ppt/presentation.xml")
    master_xml = "".join(package_text(path, name) for name in names if name.startswith("ppt/slideMasters/slideMaster") and name.endswith(".xml"))
    note_names = [name for name in names if name.startswith("ppt/notesSlides/notesSlide") and name.endswith(".xml")]
    note_text = "".join(package_text(path, name) for name in note_names)
    checks = [
        check("PPTX 投影片數", len(slide_names) == 3, "3 張 16:9 投影片"),
        check("PPTX 母片與版面", any(name.startswith("ppt/slideMasters/slideMaster") for name in names) and any(name.startswith("ppt/slideLayouts/slideLayout") for name in names), "母片與自訂版面"),
        check("PPTX 投影片編號", 'type="slidenum"' in master_xml or 'type="slidenum"' in slide_xml, "動態投影片編號欄位"),
        check("PPTX 圖表", any(name.startswith("ppt/charts/chart") for name in names) and "graphicFrame" in slide_xml, "可編輯圖表與內嵌活頁簿"),
        check("PPTX 表格", "<a:tbl>" in slide_xml, "PowerPoint 原生表格"),
        check("PPTX 圖形與排列", slide_xml.count("<p:sp>") >= 12, "多個可編輯圖形與文字物件"),
        check("PPTX 備忘稿", len(note_names) == 3 and "講者備忘稿" in note_text, "每張投影片皆有備忘稿"),
        check("PPTX 轉場", all("<p:transition" in package_text(path, name) and "<p:fade" in package_text(path, name) for name in slide_names), "每張投影片淡化轉場"),
        check("PPTX 超連結", any("hyperlink" in package_text(path, name) for name in names if name.startswith("ppt/slides/_rels/")), "外部超連結關係"),
        check("PPTX 繁中字型", "Noto Sans CJK TC" in slide_xml and "Noto Serif CJK TC" in slide_xml, "黑體與明體"),
        check("PPTX 結構順序", presentation.find("<p:notesMasterIdLst") < presentation.find("<p:sldIdLst"), "notesMasterIdLst 位於 sldIdLst 前"),
    ]
    return checks


def verify_pdfs(pdf_directory: Path) -> list[dict[str, object]]:
    checks: list[dict[str, object]] = []
    expected = {
        "OpenDeskTW_完整文字功能.pdf": ["OpenDesk TW", "第二頁用於驗證"],
        "OpenDeskTW_完整試算表功能.pdf": ["OpenDesk TW", "毛利"],
        "OpenDeskTW_完整簡報功能.pdf": ["OpenDesk TW", "資料圖表與表格"],
    }
    for filename, markers in expected.items():
        matches = list(pdf_directory.rglob(filename))
        if not matches:
            checks.append(check(f"PDF {filename}", False, "找不到輸出檔"))
            continue
        reader = PdfReader(matches[0])
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        checks.append(check(f"PDF {filename}", len(reader.pages) >= 1 and all(marker in text for marker in markers), f"{len(reader.pages)} 頁；可搜尋文字"))
    return checks


docx_path = ROOT / "OpenDeskTW_完整文字功能.docx"
xlsx_path = ROOT / "OpenDeskTW_完整試算表功能.xlsx"
pptx_path = ROOT / "OpenDeskTW_完整簡報功能.pptx"

groups = {
    "文字文件": verify_docx(docx_path),
    "試算表": verify_xlsx(xlsx_path),
    "簡報": verify_pptx(pptx_path),
}
if len(sys.argv) > 3:
    groups["PDF"] = verify_pdfs(Path(sys.argv[3]))

passed = sum(1 for checks in groups.values() for item in checks if item["passed"])
total = sum(len(checks) for checks in groups.values())
report = {
    "schemaVersion": 1,
    "suite": "OpenDesk TW Office 完整功能矩陣",
    "generatedForVersion": "2.2.2",
    "passed": passed == total,
    "summary": f"{passed}/{total} 項通過",
    "passedCount": passed,
    "totalCount": total,
    "groups": groups,
    "boundaries": [
        "VBA、ActiveX、COM、IRM 與 Microsoft 365 雲端服務不是開放格式能力，僅保留原檔並警示。",
        "複雜 SmartArt、3D 物件、PowerPoint 動作路徑與 Microsoft 專屬巨集需在實際文件逐檔確認。",
    ],
}
encoded = json.dumps(report, ensure_ascii=False, indent=2)
print(encoded)
if REPORT_PATH:
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(encoded + "\n", encoding="utf-8")
sys.exit(0 if report["passed"] else 1)
